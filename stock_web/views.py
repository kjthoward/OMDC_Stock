from django.http import HttpResponse, HttpResponseRedirect
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User, Group
from django.db.models import F, Q
from django.shortcuts import render
from django.urls import reverse
from django.db import transaction
from operator import attrgetter
import openpyxl
import pdb
import datetime
import math
import csv
import random
import string
from .prime import PRIME
from .email import send, EMAIL
from .pdf_report import report_gen
from .models import ForceReset, Suppliers, Reagents, Internal, Validation, Recipe, Inventory, Solutions, VolUsage, Projects, Storage
from .forms import LoginForm, NewInvForm1, NewInvForm, NewProbeForm, UseItemForm, OpenItemForm, ValItemForm, FinishItemForm,\
                   NewSupForm, NewReagentForm, NewRecipeForm, SearchForm, ChangeDefForm1, ChangeDefForm, RemoveSupForm,\
                   EditSupForm, EditReagForm, EditInvForm, DeleteForm, UnValForm, ChangeMinForm1, ChangeMinForm, InvReportForm,\
                   StockReportForm, PWResetForm, WitnessForm, NewProjForm, EditProjForm, ProjReportForm, RemoveProjForm, NewStoreForm, EditStoreForm, RemoveStoreForm

LOGINURL = settings.LOGIN_URL
RESETURL = "/stock/forcereset/"
UNAUTHURL = "/stock/unauth/"

CONDITIONS={"GD":"Good",
            "DU":"Damaged - Usable",
            "UN":"Damaged - Not Usable",
            "NA":"N/A, Made in House"}

#used in user_passes_test decorator to check if someone if logged in
def is_logged_in(user):
    return user.is_active
#used in user_passes_test decorator to check if the account logged in is admin
def is_admin(user):
    return user.is_staff
#used in user_passes_test decorator to check if the account has a forced password reset active (decorate used as
#even though after logging in with a reset password it prompts you to change, could go to any link manually to skip
#decorator means you will always be brought back to change password
def no_reset(user):
    if ForceReset.objects.get(user=user.pk).force_password_change==True:
        return False
    else:
        return True

#sets up database for first use (sets up groups and adds Admin as superuser/staff)
def prime(httprequest):
    if len(User.objects.all())!=0:
        messages.success(httprequest, "Database has already been primed!")
        return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        PRIME()
        messages.success(httprequest, "Database Primed")
        return HttpResponseRedirect(reverse("stock_web:listinv"))

def _toolbar(httprequest, active=""):
    inventory_dropdown = [{"name":"All", "url":reverse("stock_web:inventory", args=["_", "all","_",1])},
                          {"name":"In Stock", "url":reverse("stock_web:inventory", args=["_", "instock","_", 1])},
                          {"name":"Expiring Soon", "url":reverse("stock_web:inventory", args=["_", "expsoon","_",1])},
                          {"name":"Solutions", "url":reverse("stock_web:inventory", args=["_", "solutions","_",1])},
                          {"name":"Validated", "url":reverse("stock_web:inventory", args=["_", "validated","_",1])},
                          {"name":"Not validated", "url":reverse("stock_web:inventory", args=["_", "notvalidated","_",1])},
                          {"name":"List View","url":reverse("stock_web:listinv")},
                          ]

    toolbar = [([{"name":"Inventory", "dropdown":inventory_dropdown},
                 {"name":"Recipes", "url":reverse("stock_web:recipes"), "glyphicon":"folder-open"},
                 {"name":"Download Label File", "url":reverse("stock_web:label"), "glyphicon":"barcode"},
                 {"name":"Stock Reports", "url":reverse("stock_web:stockreport", args=["_","_"]),"glyphicon":"download"},
                 {"name":"Items to Order", "url":reverse("stock_web:toorder"),"glyphicon":"gbp"},
                 ], "left")]

    undo_dropdown = [{"name": "Change Default Supplier", "url":reverse("stock_web:changedef", args=["_"])},
                     {"name": "Edit Minimum Stock Levels", "url":reverse("stock_web:changemin",args=["_"])},
                     {"name": "(De)Activate Reagents", "url":reverse("stock_web:activreag")},
                     {"name": "(De)Activate Suppliers", "url":reverse("stock_web:activsup")},
                     {"name": "(De)Activate Projects", "url":reverse("stock_web:activproj")},
                     {"name": "(De)Activate Storage Locations", "url":reverse("stock_web:activstore")},
                     {"name": "Remove Suppliers", "url":reverse("stock_web:removesup")},
                     {"name": "Remove Project", "url":reverse("stock_web:removeproj")},
                     {"name": "Remove Storage Location", "url":reverse("stock_web:removestore")},
                     {"name": "Edit Inventory Item", "url":reverse("stock_web:editinv", args=["_"])}]


    if httprequest.user.is_staff:
        toolbar[0][0].pop()
        toolbar[0][0].pop()
        reports_dropdown=[{"name":"Inventory Reports", "url":reverse("stock_web:invreport",args=["_","_"])},
                           {"name":"Stock Reports", "url":reverse("stock_web:stockreport", args=["_","_"])},
                           {"name":"Project Reports", "url":reverse("stock_web:projreport",args=["_","_","_"])},
                           {"name":"Items to Order", "url":reverse("stock_web:toorder")},]

        toolbar[0][0].append({"name": "Reports", "glyphicon":"download", "dropdown":reports_dropdown})
        toolbar[0][0].append({"name":"Edit Data", "dropdown":undo_dropdown, "glyphicon":"wrench"})
        toolbar[0][0].append({"name":"Update Users", "url":"/stock/admin/auth/user/","glyphicon":"user"})
        new_dropdown = [{"name": "Inventory Item", "url":reverse("stock_web:newinv", args=["_"])},
                        {"name":"Supplier", "url":reverse("stock_web:newsup")},
                        {"name":"Project", "url":reverse("stock_web:newproj")},
                        {"name":"Storage Location", "url":reverse("stock_web:newstore")},
                        {"name":"Reagent", "url":reverse("stock_web:newreagent")},
                        {"name":"Recipe", "url":reverse("stock_web:newrecipe")}]
        toolbar.append(([{"name": "new", "glyphicon": "plus", "dropdown": new_dropdown}],"right"))

    else:
        toolbar.append(([{"name": "New Inventory Item", "glyphicon": "plus", "url":reverse("stock_web:newinv", args=["_"])}],"right"))


    toolbar[1][0].append({"name": "search", "glyphicon": "search", "url": reverse("stock_web:search")})
    toolbar[1][0].append({"name":"Account Settings", "glyphicon":"cog", "dropdown":[
                         {"name": "Logout "+str(httprequest.user), "url": reverse("stock_web:loginview")},
                         {"name":"Change Password", "url":reverse("stock_web:change_password")}]})

    for entry in toolbar[0][0]:
        if entry["name"] == active:
            entry["active"] = True
    for entry in toolbar[1][0]:
        if entry["name"] == active:
            entry["active"] = True
    return toolbar

@user_passes_test(is_logged_in, login_url=LOGINURL)
def change_password(httprequest):
    if httprequest.method == 'POST':
        form = PasswordChangeForm(httprequest.user, httprequest.POST)
        if form.is_valid():
            if form.cleaned_data["old_password"]==form.cleaned_data["new_password1"]:
                messages.success(httprequest, 'Your password new password cannot be your old password')
                return HttpResponseRedirect(reverse("stock_web:change_password"))
            user = form.save()
            update_session_auth_hash(httprequest, user)  # Important!
            messages.success(httprequest, 'Your password was successfully updated!')
            try:
                if ForceReset.objects.get(user=httprequest.user.pk).force_password_change==True:
                    reset=ForceReset.objects.get(user_id=user.id)
                    reset.force_password_change=False
                    reset.save()
            except:
                pass
            return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            errors=[]
            for v in form.errors.values():
                errors+=[str(v[0]).strip('<ul class="errorlist"><li><\lie></').replace("didn't", "do not")]
            messages.success(httprequest, (" ".join(errors)))
    else:
        form = PasswordChangeForm(httprequest.user)
    submiturl = reverse("stock_web:change_password")
    cancelurl = reverse("stock_web:listinv")
    context={"form": form, "heading":"Change Password for {}".format(httprequest.user),
             "submiturl": submiturl, "cancelurl": cancelurl, "toolbar":_toolbar(httprequest, active="Account Settings")}
##    if ForceReset.objects.get(user=httprequest.user.pk).force_password_change==False:
##        context.update({"toolbar":_toolbar(httprequest)})
    return render(httprequest, "stock_web/pwform.html", context)

def resetpw(httprequest):
    if EMAIL==False:
        messages.success(httprequest, "Email resets are not currently enabled. Contact an Admin to reset your password")
        return HttpResponseRedirect(reverse("stock_web:loginview"))
    if httprequest.method == 'POST':
        form=PWResetForm(httprequest.POST)

        if form.is_valid():
            new_pw=''.join(random.choice(string.ascii_letters + string.digits) for _ in range(15))
            USER = User.objects.get(username=form.data["user"])
            USER.set_password(new_pw)
            USER.save()
            try:
                reset=ForceReset.objects.get(user_id=USER.id)
                reset.force_password_change=True
            #except just incase user somehow doesn't exist in force reset table (but they should be synced)
            except:
                reset=ForceReset.objects.create(user=USER, force_password_change=True)
            reset.save()
            subject="Password for stock database account '{}' has been reset.".format(USER.username)
            text="<p>The Password for your stock database account '{}' has been reset.<br><br>".format(USER.username)
            text+="Your new password is: {}<br><br>".format(new_pw)
            text+="You will be required to change this when you login.<br><br>"
            try:
                send(subject,text,USER.email)
                messages.success(httprequest, "Password has been reset and emailed to {}".format(USER.email))
            except:
                messages.success(httprequest, "Email did not send. Please try again. If error persists contact an Admin")
            return HttpResponseRedirect(reverse("stock_web:loginview"))
    else:
        form=PWResetForm()
    submiturl = reverse("stock_web:resetpw")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/pwform.html", {"form": form, "heading":"Enter Username to Reset Password for", "submiturl": submiturl, "cancelurl": cancelurl })

def forcereset(httprequest):
    messages.success(httprequest, "You are required to change your password after resetting it")
    return HttpResponseRedirect(reverse("stock_web:change_password"))

def unauth(httprequest):
    messages.success(httprequest, "The page you are trying to access requires Admin rights.\
                     Please login using an Admin account.")
    return HttpResponseRedirect(reverse("stock_web:listinv"))

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def search(httprequest):
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
            return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = SearchForm(httprequest.POST)
            if form.is_valid():
                queries = []
                for key, query in [("reagent", "reagent__name__icontains"), ("supplier", "supplier__name__icontains"),
                                    ("project", "project__name__icontains"),("storage", "storage__name__icontains"),
                                   ("lot_no", "lot_no__icontains"), ("int_id","internal__batch_number__exact"),
                                   ("po", "po__icontains"),
                                   ("in_stock","finished__lte"),
                                  ]:
                    val = form.cleaned_data[key]
                    if val:
                        queries += ["{}={}".format(query, val)]
                return HttpResponseRedirect(reverse("stock_web:inventory", args=["search", ";".join(queries),"_","1"]))
    else:
        form = SearchForm()
    submiturl = reverse("stock_web:search")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/searchform.html", {"form": form, "heading":"Enter Search Query", "toolbar": _toolbar(httprequest, active="search"), "submiturl": submiturl, "cancelurl": cancelurl })


def loginview(httprequest):
    if httprequest.user.is_authenticated:
        logout(httprequest)
        messages.success(httprequest,"You are now logged out")
        return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:

        if httprequest.method == "POST" and "login" in httprequest.POST:
            form = LoginForm(httprequest.POST)

            if form.is_valid():
                user = authenticate(username=form.cleaned_data["username"], password=form.cleaned_data["password"])
                if user is not None and user.is_active:
                    login(httprequest, user)
                    if ForceReset.objects.get(user=httprequest.user.pk).force_password_change==True:
                        messages.success(httprequest,"You are required to change your password after resetting it")
                        return HttpResponseRedirect(reverse("stock_web:change_password"))
                    else:
                        return HttpResponseRedirect(httprequest.GET["next"] if "next" in httprequest.GET.keys() else reverse("stock_web:listinv"))

                else:
                    try:
                        User.objects.get(username=form.cleaned_data["username"])
                        messages.success(httprequest, "Incorrect password entered")
                    except:
                        messages.success(httprequest, f'User {form.cleaned_data["username"]} does not exist')
            else:
                pdb.set_trace()

        else:
            form = LoginForm()
    return render(httprequest, "stock_web/login.html", {"form": form})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def listinv(httprequest):
    title = "List of Reagents"
    headings = ["Reagent Name", "Number In Stock", "Minimum Stock Level"]
    items=Reagents.objects.all().exclude(is_active=False, count_no=0).order_by("name")
    body=[]
    for item in items:
        values = [item.name,
                  "{}µl".format(item.count_no) if item.track_vol==True else item.count_no,
                  "{}µl".format(item.min_count) if item.track_vol==True else item.min_count]
        urls=[reverse("stock_web:inventory",args=["filter","reagent__name__iexact={};finished__lte=0".format(item.name),"_",1]),
              "",
              "",
              ]
        body.append((zip(values,urls), True if item.count_no<item.min_count else False))

    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest, active="Inventory")}
    return render(httprequest, "stock_web/list.html", context)

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def inventory(httprequest, search, what, sortby, page):
    #forces page 1 if non numberical value entered or <1
    try:
        page=int(page)
    except:
        page=1
    if page<1:
        page=1
    ##HACK USED IF / IS IN REAGENT SEARCH TERM (such as "item1/item2 mix")
    if "/" in search and (("__icontains" in what) or ("__lte" in what) or ("__iexact" in what)):
        what="".join(search.split("/")[1:])+"/"+what
        search="search"
    if sortby!="_":
        sortquery = sortby.split("=")[1]
        if sortquery=="days_rem":
            pos=True
            sortby="_"
        elif sortquery=="-days_rem":
            pos=False
            sortby="_"
    else:
        sortquery="_"
    if what=="all":
        title = "Inventory - All Items"
        if sortby!="_":
            items=Inventory.objects.all().order_by(sortquery)
        else:
            items=Inventory.objects.all()
    if what=="instock":
        title = "Inventory - Items In Stock"
        if sortby!="_":
            items=Inventory.objects.filter(finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.filter(finished=False)
    elif what=="solutions":
        title = "Inventory - Solutions"
        if sortby!="_":
            items=Inventory.objects.filter(sol_id__isnull=False, finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.filter(sol_id__isnull=False, finished=False)
    elif what=="validated":
        title = "Inventory - Validated Items"
        if sortby!="_":
            items=Inventory.objects.filter(val_id__isnull=False, finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.filter(val_id__isnull=False, finished=False)
    elif what=="notvalidated":
        title = "Inventory - Items Not Validated"
        if sortby!="_":
            items=Inventory.objects.filter(val_id__isnull=True, finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.filter(val_id__isnull=True, finished=False)
    elif what=="expsoon":
        title = "Inventory - Items Expiring Within 6 Weeks"
        if sortby!="_":
            items=Inventory.objects.filter(date_exp__lte=datetime.datetime.now()+datetime.timedelta(days=42), finished=False).order_by(sortquery)
        else:
            items=Inventory.objects.filter(date_exp__lte=datetime.datetime.now()+datetime.timedelta(days=42), finished=False)
    elif search=="search" or search=="filter":
        query = dict([q.split("=") for q in what.split(";")])
        if search=="search":
            title="Search Results"
        elif search=="filter" and "reagent__name__iexact" in query.keys():
            title=query['reagent__name__iexact']

        if sortby!="_":
            items=Inventory.objects.filter(**query).order_by(sortquery)
        else:
            items=Inventory.objects.filter(**query)
        if len(items)==1:
            return HttpResponseRedirect(reverse("stock_web:item",args=[items[0].id]))
    items=items.select_related("supplier","reagent","internal","project","project_used")
    pages=[]
    if len(items)>200:
        for i in range(1, math.ceil(len(items)/200)+1):
            pages+=[[i,reverse("stock_web:inventory", args=[search, what, sortby, i]) if page!=i else ""]]
        #forces go to page 1 if number>last page manually entered
        if page>pages[-1][0]:
             return HttpResponseRedirect(reverse("stock_web:inventory", args=[search, what, sortby, 1]))
    headings = ["Reagent Name", "Supplier", "Batch ID", "Date Received", "Expiry Date", "Project", "Storage Location", "Date Opened", "Days till expired"]
    headurls = [reverse("stock_web:inventory", args=[search, what,"order=-reagent_id__name"
                                                     if sortby=="order=reagent_id__name" else "order=reagent_id__name", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-supplier_id__name"
                                                     if sortby=="order=supplier_id__name" else "order=supplier_id__name", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-internal_id__batch_number"
                                                     if sortby=="order=internal_id__batch_number" else "order=internal_id__batch_number", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_rec"
                                                     if sortby=="order=date_rec" else "order=date_rec", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_exp"
                                                     if sortby=="order=date_exp" else "order=date_exp", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-project_id__name"
                                                     if sortby=="order=project_id__name" else "order=project_id__name", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-storage_id__name"
                                                     if sortby=="order=storage_id__name" else "order=storage_id__name", 1]),
                reverse("stock_web:inventory", args=[search, what,"order=-date_op"
                                                     if sortby=="order=date_op" else "order=date_op",1]),
                reverse("stock_web:inventory", args=[search, what,"order=-days_rem"
                                                     if sortquery=="days_rem" else "order=days_rem",1])]
    headings=zip(headings,headurls)
    body=[]

    if "date_op" in sortby:
        q=items.extra(select={'date_op_null': 'date_op is null'})
        items = q.extra(order_by=['date_op_null',sortquery])
        ####################################################################
    if "days_rem" in sortquery:
        if pos==True:
            items = sorted(items, key = lambda item:int(item.days_remaining()), reverse=False)
        elif pos==False:
            items = sorted(items, key = lambda item:int(item.days_remaining()), reverse=True)



    items_trunc=items[(page-1)*200:page*200]
    for item in items_trunc:
        values = [item.reagent.name,
                  item.supplier.name,
                  item.internal.batch_number,
                  item.date_rec.strftime("%d/%m/%y"),
                  item.date_exp.strftime("%d/%m/%y"),
                  item.project.name if item.project is not None else "",
                  item.storage.name if item.storage is not None else "",
                  item.date_op.strftime("%d/%m/%y") if item.date_op is not None else "",
                  item.days_remaining(),
                  ]
        urls=[reverse("stock_web:item",args=[item.id]),
              "",
              "",
              "",
              "",
              "",
              "",
              "",
              "",
              ]
        body.append((zip(values,urls),item.finished))

    context = {"header":title,"headings":headings, "body":body,
               "toolbar":_toolbar(httprequest, active="Inventory")}
    if pages:
        context.update({"pages":pages, "text1": "Click to change page",
                        "text2":"Current page is {} showing items {}-{}".format(page,(page-1)*200,
                                                                         page*200 if (page*200<len(items)) else (len(items)))})

    return render(httprequest, "stock_web/listinventory.html", context)

@user_passes_test(is_logged_in, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def toorder(httprequest):
    low_reagents=Reagents.objects.filter(count_no__lt=F("min_count")).exclude(name="INTERNAL").exclude(is_active=False)
    if len(low_reagents)==0:
        messages.success(httprequest, "No items are in need of ordering")
        return HttpResponseRedirect(httprequest.META.get('HTTP_REFERER', '/'))
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Items_To_Order_Report_{}.csv"'.format(str(datetime.datetime.today().strftime("%d/%m/%Y")))
        writer = csv.writer(response)
        writer.writerow(["Item", "Project Last Used By", "Default Supplier", "Default Supplier Catalogue Number", "Amount to Order"])
        for item in low_reagents:
            open_list=Inventory.objects.filter(reagent=item).order_by("date_op").reverse()
            if len(open_list)!=0:
                last_open=open_list[0]
                try:
                    project_last_used=last_open.project_used.name
                except:
                    project_last_used=None
                if project_last_used==None:
                    project_last_used="N/A"
            else:
                project_last_used="N/A"
            writer.writerow([item.name, project_last_used, item.supplier_def.name, item.cat_no, int(item.min_count)-int(item.count_no)])
        return response

@user_passes_test(is_logged_in, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def stockreport(httprequest, pk, extension):
    submiturl = reverse("stock_web:stockreport",args=[pk,extension])
    cancelurl = reverse("stock_web:listinv")
    if httprequest.user.is_staff==True:
        toolbar = _toolbar(httprequest, active="Reports")
    else:
        toolbar = _toolbar(httprequest, active="Stock Reports")
    header = "Select Reagent to Generate Stock Report For"
    form=StockReportForm
    if pk=="_":
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if "pdf" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:stockreport", args=[form.cleaned_data["name"].pk,0]))
                    elif "xlsx" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:stockreport", args=[form.cleaned_data["name"].pk,1]))
        else:
            form = form()

    else:
        title="{} - Stock Report".format(Reagents.objects.get(pk=int(pk)))
        #gets items, with open items first, then sorted by expirey date
        items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user","project", "project_used").filter(reagent_id=int(pk),finished=False).order_by("-is_op","date_exp")
        body=[["Supplier Name", "Lot Number", "Project Assigned", "Used by Project", "Stock Number", "Date Received",
               "Expiry Date", "Date Open", "Opened By", "Date Validated", "Validation Run"]]

        for item in items:
            body+= [[ item.supplier.name,
                      item.lot_no,
                      item.project.name if item.project is not None else "",
                      item.project_used.name if item.project_used_id is not None else "",
                      item.internal.batch_number,
                      item.date_rec.strftime("%d/%m/%y"),
                      item.date_exp.strftime("%d/%m/%y"),
                      item.date_op.strftime("%d/%m/%y") if item.date_op is not None else "",
                      item.op_user.username if item.op_user is not None else "",
                      item.val.val_date.strftime("%d/%m/%y") if item.val is not None else "",
                      item.val.val_run if item.val is not None else "",
                      ]]
        if extension=='0':
            httpresponse = HttpResponse(content_type='application/pdf')
            httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(title)
            table=report_gen(body,title,httpresponse,httprequest.user.username)

        if extension=='1':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row in body:
                worksheet.append(row)
            httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
            httpresponse['Content-Disposition'] = 'attachment; filename={}.xlsx'.format("_".join(title.split()))
        return httpresponse
    return render(httprequest, "stock_web/reportform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def invreport(httprequest,what, extension):
    submiturl = reverse("stock_web:invreport",args=[what,extension])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Reports")
    header = "Select Inventory Report To Generate"
    form=InvReportForm
    if what=="_":
        if httprequest.method=="POST":

            if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():

                    if "pdf" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:invreport", args=[form.cleaned_data["report"],0]))
                    elif "xlsx" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:invreport", args=[form.cleaned_data["report"],1]))
        else:
            form = form()
    else:
        if what=="unval":
            title="All Unvalidated Items Report"
            items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user","project_used", "project").filter(val_id=None,finished=False).order_by("reagent_id__name","-is_op","date_exp")
        elif what=="val":
            title="All Validated Items Report"
            items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user","project_used", "project").filter(val_id__gte=0,finished=False).order_by("reagent_id__name","-is_op","date_exp")
        elif what=="exp":
            title="Items Expiring Soon Report"
            items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user","project_used", "project").filter(date_exp__lte=datetime.datetime.now()+datetime.timedelta(days=42),finished=False).order_by("reagent_id__name","-is_op","date_exp")
        elif what=="all":
            title="All Items In Stock Report"
            items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user","project_used", "project").filter(is_op=False,finished=False).order_by("reagent_id__name","-is_op","date_exp")
        elif what=="allinc":
            title="All Items In Stock Including Open Report"
            items = Inventory.objects.select_related("supplier","reagent","internal","val","op_user","project_used", "project").filter(finished=False).order_by("reagent_id__name","-is_op","date_exp")

        if what=="all":
            body=[["Reagent", "Supplier", "Project", "Used by Project", "Lot Number", "Stock Number", "Received",
                   "Expiry"]]
            for item in items:
                body+= [[item.reagent.name,
                          item.supplier.name,
                          item.project.name if item.project_id is not None else "",
                          item.project_used.name if item.project_used_id is not None else "",
                          item.lot_no,
                          item.internal.batch_number,
                          item.date_rec.strftime("%d/%m/%y"),
                          item.date_exp.strftime("%d/%m/%y"),
                          ]]
        elif what=="unval":
            body=[["Reagent", "Supplier", "Lot Number", "Project", "Used by Project", "Stock Number", "Received",
                   "Expiry", "Opened", "Opened By"]]
            for item in items:
                body+= [[item.reagent.name,
                          item.supplier.name,
                          item.lot_no,
                          item.project.name if item.project is not None else "",
                          item.project_used.name if item.project_used_id is not None else "",
                          item.internal.batch_number,
                          item.date_rec.strftime("%d/%m/%y"),
                          item.date_exp.strftime("%d/%m/%y"),
                          item.date_op.strftime("%d/%m/%y") if item.date_op is not None else "",
                          item.op_user.username if item.op_user is not None else "",
                          ]]

        else:
            body=[["Reagent", "Supplier", "Lot Number", "Project", "Used by Project", "Stock Number", "Received",
                   "Expiry", "Opened", "Opened By", "Date Validated", "Validation Run"]]
            for item in items:
                body+= [[item.reagent.name,
                          item.supplier.name,
                          item.lot_no,
                          item.project.name if item.project is not None else "",
                          item.project_used.name if item.project_used_id is not None else "",
                          item.internal.batch_number,
                          item.date_rec.strftime("%d/%m/%y"),
                          item.date_exp.strftime("%d/%m/%y"),
                          item.date_op.strftime("%d/%m/%y") if item.date_op is not None else "",
                          item.op_user.username if item.op_user is not None else "",
                          item.val.val_date.strftime("%d/%m/%y") if item.val is not None else "",
                          item.val.val_run if item.val is not None else "",
                          ]]
        if extension=='0':
            httpresponse = HttpResponse(content_type='application/pdf')
            httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(title)
            table=report_gen(body,title,httpresponse,httprequest.user.username)

        if extension=='1':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row in body:
                worksheet.append(row)
            httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
            httpresponse['Content-Disposition'] = 'attachment; filename={}.xlsx'.format("_".join(title.split()))
        return httpresponse
    return render(httprequest, "stock_web/reportform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def label(httprequest):
    submiturl = reverse("stock_web:listinv")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Download Label File")
    header = "Press 'Download' to download the Excel file for all new items"
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            items = Inventory.objects.select_related("project", "project_used", "reagent","internal").filter(printed=False)
            if len(items)==0:
                messages.success(httprequest, "No items exist that are marked as unprinted")
                return HttpResponseRedirect(reverse("stock_web:label"))
            else:
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                for item in items:
                    worksheet.append([item.reagent.name, item.project.name if item.project is not None else "OMDC", item.date_rec, item.internal.batch_number])
                    item.printed=True
                    item.save()
                httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
                httpresponse['Content-Disposition'] = 'attachment; filename="Stock_Label_File - {}.xlsx"'.format(str(datetime.datetime.today().strftime("%d/%m/%Y")))

                #OLD CSV, Keeping incase need to change back
                # response = HttpResponse(content_type='text/csv')
                # response['Content-Disposition'] = 'attachment; filename="Stock_Label_File - {}.csv"'.format(str(datetime.datetime.today().strftime("%d/%m/%Y")))
                # writer = csv.writer(response)
                # for item in items:
                #     writer.writerow([item.reagent.name, item.internal,item.project if item.project is not None else "OMDC"])
                #     item.printed=True
                #     item.save()
                return httpresponse
    else:
        pass
    return render(httprequest, "stock_web/labelform.html", {"header": header, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def projreport(httprequest, pk, extension, fin):
    submiturl = reverse("stock_web:projreport",args=[pk,extension, fin])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Reports")
    header = "Select Project to Generate Stock Report For"
    form=ProjReportForm
    if pk=="_":
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or "Download" not in httprequest.POST["submit"]:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if "pdf" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:projreport", args=[form.cleaned_data["name"].pk,0, form.cleaned_data["in_stock"]]))
                    elif "xlsx" in httprequest.POST["submit"]:
                        return HttpResponseRedirect(reverse("stock_web:projreport", args=[form.cleaned_data["name"].pk,1, form.cleaned_data["in_stock"]]))
        else:
            form = form()

    else:
        title="{} - Project Stock Report".format(Projects.objects.get(pk=int(pk)))
        #gets items, with open items first, then sorted by expirey date
        items = Inventory.objects.select_related("supplier","reagent", "project_used", "project", "internal","val","op_user").filter(project_id=int(pk))
        if fin=="0":
            items=items.exclude(finished=True)
        items.order_by("-finished","-is_op","date_exp")
        body=[["Reagent", "Supplier Name", "Lot Number", "Date Received",
               "Expiry Date", "Date Open", "Opened By", "Project Used", "Date Validated", "Validation Run", "Finished"]]

        for item in items:
            body+= [[ item.reagent.name,
                      item.supplier.name,
                      item.lot_no,
                      item.date_rec.strftime("%d/%m/%y"),
                      item.date_exp.strftime("%d/%m/%y"),
                      item.date_op.strftime("%d/%m/%y") if item.date_op is not None else "",
                      item.op_user.username if item.op_user is not None else "",
                      item.project_used.name if item.project_used_id is not None else "",
                      item.val.val_date.strftime("%d/%m/%y") if item.val is not None else "",
                      item.val.val_run if item.val is not None else "",
                      item.finished,
                      ]]
        if extension=='0':
            httpresponse = HttpResponse(content_type='application/pdf')
            httpresponse['Content-Disposition'] = 'attachment; filename="{}.pdf"'.format(title)
            table=report_gen(body,title,httpresponse,httprequest.user.username)

        if extension=='1':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row in body:
                worksheet.append(row)
            httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
            httpresponse['Content-Disposition'] = 'attachment; filename={}.xlsx'.format("_".join(title.split()))
        return httpresponse
    return render(httprequest, "stock_web/reportform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})


def _item_context(httprequest, item, undo):
    title = ["Reagent -",
             "Supplier -",
             "Lot Number -",
             "Stock Number -"]
    title_values=[item.reagent.name,
                  item.supplier.name,
                  item.lot_no if item.lot_no else "",
                  item.internal.batch_number]
    title_url=["","","",""]
    if undo=="undo":
        title[0:0]=["***WARNING - ONLY TO BE USED TO CORRECT DATA ENTRY ERRORS. IT MAY NOT BE POSSIBLE TO UNDO CHANGES MADE HERE***"]
        title_values.append("")
        title_url.append("")
    if item.project_id is not None:
        title.append("Project -")
        title_values.append(item.project.name)
        title_url.append("")
    if item.project_used_id is not None:
        title.append("Used By Project -")
        title_values.append(item.project_used.name)
        title_url.append("")
    if item.storage is not None:
        title.append("Location -")
        title_values.append(item.storage.name)
        title_url.append("")
    if item.po is not None:
        title.append("Purchase Order Number -")
        title_values.append(item.po)
        title_url.append("")
    if item.sol is not None and undo!="undo":
        title.append("Witnessed By -")
        title_values.append(item.witness)
        title_url.append("")
        for comp in item.sol.list_comp():
            title.append("")
            title_values.append(comp)
            title_url.append(reverse("stock_web:item",args=[comp.id]))

    title=zip(title, title_values ,title_url)
    if item.sol is not None:
        headings = ["Date Created", "Created By", "Condition Received", "Expiry Date"]
    else:
        headings = ["Date Received", "Received By", "Condition Received", "Expiry Date"]
    values = [item.date_rec.strftime("%d/%m/%y"), item.rec_user.username, CONDITIONS[item.cond_rec], item.date_exp]
    urls = ["", "", "", ""]
    SKIP=False
    if item.date_op is not None:
        headings+=["Date Opened", "Opened By"]
        values+=[item.date_op,item.op_user]
        urls+=["",""]
    if item.val_id is not None:
        headings+=["Date Validated", "Validation Run", "Validation User"]
        values+=[item.val.val_date, item.val.val_run, item.val.val_user]
        urls+=["","",""]
    if ((item.date_op is None) and (item.finished==False)):
        headings+=["Action"]
        if undo=="undo":
            values+=["Delete Item"]
            urls+=[reverse("stock_web:undoitem",args=["delete",item.id])]
        else:
            values+=["Open Item"]
            urls+=[reverse("stock_web:openitem",args=[item.id])]
    if ((item.date_op is not None) and (item.val_id is None) and (item.finished==False)):
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-open Item"]
            urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
        else:
            headings+=["Action"]
            values+=["Validate Item"]
            urls+=[reverse("stock_web:valitem",args=[item.id])]
    elif ((item.date_op is not None) and (item.val_id is not None) and (item.finished==False) and (undo=="undo")):
        headings+=["Action"]
        values+=["Un-open Item"]
        urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
    if item.val_id is not None and item.finished==False:
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-Validate Item"]
            urls+=[reverse("stock_web:undoitem",args=["unval",item.id])]
        elif item.is_op==True:
            headings+=["Action"]
            values+=["Finish/Discard Item"]
            urls+=[reverse("stock_web:finishitem",args=[item.id])]
            SKIP=True
    if ((item.finished==False) and (undo!="undo")  and (SKIP==False)):
        headings+=["Action"]
        values+=["Discard Item"]
        urls+=[reverse("stock_web:finishitem",args=[item.id])]
    if item.finished==True:
        if item.is_op==True:
            headings+=["Date Finished", "Finished by"]
        else:
            headings+=["Date Discarded", "Discared by"]
        values+=[item.date_fin, item.fin_user]
        urls+=["",""]
        if undo=="undo":
            headings+=["Action"]
            values+=["Re-Open Item"]
            urls+=[reverse("stock_web:undoitem",args=["reopen",item.id])]
    body = [(zip(values,urls, urls),False)]
    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest), "track_vol":False, "label": item.printed}
    if ((item.finished==True) and (item.fin_text is not None)):
        context.update({"newinformation":item.fin_text})
    return context

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def _vol_context(httprequest, item, undo):
    stripe=False

    title = ["Reagent -",
             "Supplier -",
             "Purchase Order Number -",
             "Lot Number -",
             "Stock Number -",
             "Volume Received -" if item.sol is None else "Volume Made Up -",
             "Current Volume - "]
    title_values=[item.reagent.name,
                  item.supplier.name,
                  item.po,
                  item.lot_no if item.lot_no else "",
                  item.internal.batch_number,
                  item.vol_rec,
                  item.current_vol if item.current_vol is not None else 0]
    title_url=["","","","","","","",""]
    skip=False
    if undo=="undo":
        title[0:0]=["***WARNING - ONLY TO BE USED TO CORRECT DATA ENTRY ERRORS. IT MAY NOT BE POSSIBLE TO UNDO CHANGES MADE HERE***"]
        title_values.append("")
        title_url.append("")
    if item.project_id is not None:
        title.append("Project -")
        title_values.append(item.project.name)
        title_url.append("")
    if item.project_used_id is not None:
        title.append("Used By Project -")
        title_values.append(item.project_used.name)
        title_url.append("")
    if item.storage is not None:
        title.append("Location -")
        title_values.append(item.storage.name)
        title_url.append("")
    if item.sol is not None:
        title.append("Witnessed By -")
        title_values.append(item.witness)
        title_url.append("")
        for comp in item.sol.list_comp():
            title.append(comp)
            title_values.append("")
            title_url.append(reverse("stock_web:item",args=[comp.id]))

    title=zip(title,title_values,title_url)
    if item.sol is not None:
        headings = ["Date Created", "Created By", "Condition Received", "Expiry Date"]
    else:
        headings = ["Date Received", "Received By", "Condition Received", "Expiry Date"]
    values = [item.date_rec.strftime("%d/%m/%y"), item.rec_user.username, CONDITIONS[item.cond_rec], item.date_exp]
    urls = ["", "", "", ""]
    SKIP=False
    if item.date_op is not None:
        headings+=["Date Opened", "Opened By"]
        values+=[item.date_op,item.op_user]
        urls+=["",""]
    if item.val_id is not None:
        headings+=["Date Validated", "Validation Run", "Validation User"]
        values+=[item.val.val_date, item.val.val_run, item.val.val_user]
        urls+=["","",""]
    if ((item.date_op is None) and (item.finished==False)):
        headings+=["Action"]
        if undo=="undo":
            values+=["Delete Item"]
            urls+=[reverse("stock_web:undoitem",args=["delete",item.id])]
        else:
            values+=["Open Item"]
            urls+=[reverse("stock_web:openitem",args=[item.id])]
    if ((item.is_op==True) and (item.finished==False) and (undo!="undo")):
        headings+=["Action"]
        values+=["Use Amount"]
        urls+=[reverse("stock_web:useitem",args=[item.id])]
    if ((item.date_op is not None) and (item.val_id is None) and (item.finished==False)):
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-open Item"]
            urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
        elif item.last_usage is not None:
            headings+=["Action"]
            values+=["Validate Item"]
            urls+=[reverse("stock_web:valitem",args=[item.id])]
    elif ((item.date_op is not None) and (item.val_id is not None) and (item.finished==False) and (undo=="undo")):
        headings+=["Action"]
        values+=["Un-open Item"]
        urls+=[reverse("stock_web:undoitem",args=["unopen",item.id])]
    if item.val_id is not None and item.finished==False:
        if undo=="undo":
            headings+=["Action"]
            values+=["Un-Validate Item"]
            urls+=[reverse("stock_web:undoitem",args=["unval",item.id])]
        elif item.is_op==True:
            headings+=["Action"]
            values+=["Finish/Discard Item"]
            urls+=[reverse("stock_web:finishitem",args=[item.id])]
            SKIP=True
    if ((item.finished==False) and (undo!="undo")  and (SKIP==False)):
        headings+=["Action"]
        values+=["Discard Item"]
        urls+=[reverse("stock_web:finishitem",args=[item.id])]
    if item.finished==True:
        if item.is_op==True:
            headings+=["Date Finished", "Finished by"]
        else:
            headings+=["Date Discarded", "Discared by"]
        values+=[item.date_fin, item.fin_user]
        urls+=["",""]
    body = [(zip(values,urls, urls),stripe)]
    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest),"label":item.printed}
    if ((item.finished==True) and (item.fin_text is not None)):
        context.update({"newinformation":item.fin_text})
    if item.current_vol<item.vol_rec:
        vol_headings=["Volume at Start", "Volume at End", "Volume Used", "Date", "User"]
        if undo=="undo":
            vol_headings+=["Action"]
        uses=VolUsage.objects.filter(item=item.pk)
        uses=sorted(uses, key = lambda use:use.date)
        vol_body=[]
        for use in uses:
            values=[use.start,
                    use.end,
                    use.used,
                    use.date,
                    use.user]
            urls=["","","","",""]
            style=["","","","",""]
            if use.sol is not None:
                urls=[reverse("stock_web:item",args=[Inventory.objects.get(sol=use.sol).pk])]*5


            if undo=="undo":
                if use==item.last_usage:
                    if item.finished==True:
                        values+=["UNDO (RE-OPEN)"]
                    else:
                        values+=["UNDO"]
                    urls+=[reverse("stock_web:undoitem",args=["unuse",item.id])]
                    style+=[""]
                else:
                    values+=[""]
                    urls+=[""]
                    style+=[""]
            vol_body.append((zip(values,urls,urls),stripe))
            stripe=not(stripe)
        context.update({"track_vol":True,
                        "vol_headings":vol_headings,
                        "vol_body":vol_body})
    return context

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def useitem(httprequest,pk):
    item=Inventory.objects.get(pk=int(pk))
    if item.is_op==False:
        return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    uses=VolUsage.objects.filter(item=item)
    if item.is_op==True and item.val is None and len(uses)>0 and item.sol is None:
        messages.success(httprequest, "WARNING - ITEM IS NOT VALIDATED")
    form=UseItemForm
    header=["Enter Volume used for {} - Current Volume is {}µl".format(item, item.current_vol)]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:

            if form.is_valid():
                Inventory.take_out(form.cleaned_data["vol_used"], int(pk), httprequest.user, form.cleaned_data["date_used"])
                item.refresh_from_db()
                message=[]
                if item.reagent.count_no<item.reagent.min_count:
                    if item.sol is not None:
                        make="made"
                    else:
                        make="ordered"
                    message+=["Current stock level for {} is {}µl. Minimum quantity is {}µl. Check if more needs to be {}".format(item.reagent.name,item.reagent.count_no,
                                                                                                                          item.reagent.min_count, make)]


                    if EMAIL==True:
                        subject="{} - Stock Level is below minimum level".format(item.reagent.name)
                        text="<p>Item {} (Catalogue number {}) has a stock level of {}µl.<br><br>".format(item.reagent.name, item.reagent.cat_no, item.reagent.count_no)
                        text+="The last project to use this item was {}.<br><br>".format(item.project.name if item.project is not None else 'NOT ASSIGNED')
                        text+="Minimum Stock level for this item is {}µl.<br><br>".format(item.reagent.min_count)
                        for user in User.objects.filter(is_staff=True, is_active=True):
                            if user.email!="":
                                try:
                                    send(subject,text, user.email)
                                except Exception as e:
                                    print(e)
                if int(item.current_vol)==0:
                    message+=["THIS TUBE IS EMPTY, PLEASE DISCARD IT!"]
                if message!=[]:
                    messages.success(httprequest," ".join(message))

                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        form=form(instance=item,initial = {"date_used":datetime.datetime.now()})
    submiturl = reverse("stock_web:useitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})


@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def openitem(httprequest, pk):
    item=Inventory.objects.get(pk=int(pk))
    form=OpenItemForm
    header=["Opening item {}".format(item)]
    header+=["Date Received: {}".format(item.date_rec.strftime("%d/%m/%y"))]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            if form.is_valid():
                Inventory.open(form.cleaned_data, pk, httprequest.user)
                item.refresh_from_db()
                if item.reagent.track_vol==False:
                    if item.reagent.count_no<item.reagent.min_count:
                        if item.sol is not None:
                            make="made"
                        else:
                            make="ordered"
                        messages.success(httprequest, "Current stock level for {} is {}. Minimum quantity is {}. Check if more needs to be {}".format(item.reagent.name,
                                                                                                                                              item.reagent.count_no,
                                                                                                                                              item.reagent.min_count,
                                                                                                                                              make))
                        if EMAIL==True:
                            subject="{} - Stock Level is below minimum level".format(item.reagent.name)
                            text="<p>Item {} (Catalogue Number {}) has a stock level of {}.<br><br>".format(item.reagent.name, item.reagent.cat_no, item.reagent.count_no)
                            text+="The last project to use this item was {}.<br><br>".format(item.project.name if item.project is not None else 'NOT ASSIGNED')
                            text+="Minimum Stock level for this item is {}.<br><br>".format(item.reagent.min_count)
                            for user in User.objects.filter(is_staff=True, is_active=True):
                                if user.email!="":
                                    try:
                                        send(subject,text, user.email)
                                    except Exception as e:
                                        print(e)
                #Shows a warning if the item is opened after it's expiry date
                if form.cleaned_data["date_op"]>=item.date_exp:
                    messages.success(httprequest,"WARNING - ITEM OPEN AFTER EXPIRY DATE")
                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        if item.is_op==True:
            return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
        form=form(instance=item,initial = {"date_op":datetime.datetime.now(),"project":item.project})
    submiturl = reverse("stock_web:openitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def valitem(httprequest,pk):
    item=Inventory.objects.get(pk=int(pk))
    form=ValItemForm
    if item.is_op==False:
        return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    header=["Validating item {}".format(item)]
    header+=["Date Open: {}".format(item.date_op.strftime("%d/%m/%y"))]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            if form.is_valid():
                Inventory.validate(form.cleaned_data, item, item.lot_no, item.date_rec, httprequest.user)
                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        if Inventory.objects.get(pk=int(pk)).val is not None:
            return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
        form=form(instance=item,initial = {"val_date":datetime.datetime.now()})
    submiturl = reverse("stock_web:valitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def finishitem(httprequest, pk):
    item=Inventory.objects.get(pk=int(pk))
    form=FinishItemForm
    header=["Finishing item {}".format(item)]
    header+=["Date Open: {}".format(item.date_op.strftime("%d/%m/%y") if item.is_op==True else "NOT OPEN")]
    header+=["Date Validated: {}".format(item.val.val_date.strftime("%d/%m/%y") if item.val is not None else "NOT VALIDATED")]
    if httprequest.method=="POST":
        form = form(httprequest.POST, instance=item)
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            if form.is_valid():

                Inventory.finish(form.cleaned_data, pk, httprequest.user)
                if item.reagent.track_vol==True or item.is_op==False:
                    if item.reagent.count_no<item.reagent.min_count:
                        if item.sol is not None:
                            make="made"
                        else:
                            make="ordered"
                        messages.success(httprequest, "Current stock level for {0} is {1}{2}. Minimum quantity is {3}{2}. Check if more needs to be {4}".format(item.reagent.name,
                                                                                                                                              item.reagent.count_no,
                                                                                                                                              "µl" if item.reagent.track_vol else "",
                                                                                                                                              item.reagent.min_count,
                                                                                                                                              make))
                        if EMAIL==True:
                            subject="{} - Stock Level is below minimum level".format(item.reagent.name)
                            text="<p>Item {} (Catalogue Number {}) has a stock level of {}.<br><br>".format(item.reagent.name,item.reagent.cat_no, item.reagent.count_no)
                            text+="The last project to use this item was {}.<br><br>".format(item.project.name if item.project is not None else 'NOT ASSIGNED')
                            text+="\n\nMinimum Stock level for this item is {}.<br><br>".format(item.reagent.min_count)
                            for user in User.objects.filter(is_staff=True, is_active=True):
                                if user.email!="":
                                    try:
                                        send(subject,text, user.email)
                                    except Exception as e:
                                        print(e)

                if item.val_id is None and item.is_op==True and item.sol is None:
                    if EMAIL==True:
                        subject="{} - Discarded without validation".format(item.reagent.name)
                        text="<p>Item {} ({}) has been discarded by {} without having validation data.<br><br>".format(item.reagent.name,item.internal.batch_number,httprequest.user.username)
                        text+="This item was from the project: {}.<br><br>".format(item.project.name if item.project is not None else 'NOT ASSIGNED')
                        text+="\n\nThe reason they entered was: '{}'<br><br>".format(form.cleaned_data["fin_text"] if form.cleaned_data["fin_text"]!=None else "NOT ENTERED")
                        for user in User.objects.filter(is_staff=True, is_active=True):
                            if user.email!="":
                                try:
                                    send(subject,text, user.email)
                                except Exception as e:
                                    print(e)
                return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
    else:
        if Inventory.objects.get(pk=int(pk)).finished==True:
            return HttpResponseRedirect(reverse("stock_web:item",args=[pk]))
        if item.is_op==False:
            messages.success(httprequest,"WARNING - ITEM HAS NOT BEEN OPENED")
        if item.val_id is None and item.is_op==True and item.sol is None:
            messages.success(httprequest,"WARNING - THIS ITEM HAS NOT BEEN VALIDATED")
        form=form(instance=item,initial = {"date_fin":datetime.datetime.now(),
                                           "date_rec":item.date_rec})
    submiturl = reverse("stock_web:finishitem",args=[pk])
    cancelurl = reverse("stock_web:item",args=[pk])
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": _toolbar(httprequest), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def item(httprequest, pk):
    item = Inventory.objects.select_related("supplier","reagent", "project_used", "project", "internal","val","project").get(pk=int(pk))
    if httprequest.method=="POST":

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append([item.reagent.name, item.project.name if item.project is not None else "OMDC", item.date_rec, item.internal.batch_number])
        httpresponse = HttpResponse(content=openpyxl.writer.excel.save_virtual_workbook(workbook), content_type='application/ms-excel')
        httpresponse['Content-Disposition'] = 'attachment; filename="{}_Label_File - {}.xlsx"'.format(str(item),str(datetime.datetime.today().strftime("%d/%m/%Y")))

        #old csv label style
        # response = HttpResponse(content_type='text/csv')
        # response['Content-Disposition'] = 'attachment; filename="{}_Label_File - {}.csv"'.format(str(item),str(datetime.datetime.today().strftime("%d/%m/%Y")))
        # writer = csv.writer(response)
        # writer.writerow([item.reagent.name, item.internal,item.project if item.project is not None else "OMDC"])
        return httpresponse
    if item.reagent.track_vol==False:
        return render(httprequest, "stock_web/list_item.html", _item_context(httprequest, item, "_"))
    else:
        return render(httprequest, "stock_web/list_item.html", _vol_context(httprequest, item, "_"))

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def recipes(httprequest):
    title = "List of Recipes"
    headings = ["Recipe Name", "Number of Components", "Shelf Life (Months)", "Active"]
    items=Recipe.objects.all().order_by("name")
    body=[]

    for item in items:
        values = [item.name,
                  item.length,
                  item.shelf_life,
                  "YES" if Reagents.objects.get(recipe=item).is_active else "NO"]
        urls=[reverse("stock_web:recipe",args=[item.id]),
              "",
              "",
              "",
              ]
        body.append((zip(values,urls), False))

    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest, active="Recipes")}
    return render(httprequest, "stock_web/list.html", context)

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def recipe(httprequest, pk):
    item = Recipe.objects.get(pk=int(pk))
    title = "Components for {}".format(item.name)
    headings = ["Reagent", "Amount in stock"]
    body=[]
    for i in range(1, item.length()+1):
        values = [eval('item.comp{}.name'.format(i)),
                  "{}µl".format(eval('item.comp{}.count_no'.format(i))) if eval('item.comp{}.track_vol'.format(i))==True else eval('item.comp{}.count_no'.format(i)),
                  ]
        urls= ["",
               ""]
        body.append((zip(values,urls),False))
    context = {"header":title,"headings":headings, "body":body, "toolbar":_toolbar(httprequest, active="Recipe")}
    return render(httprequest, "stock_web/list.html", context)

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newinv(httprequest, pk):
    if pk=="_":
        title="Select Reagent to Book-in"
        template="stock_web/invform.html"
        form=NewInvForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "book-in":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if form.cleaned_data["reagent"].recipe is None:
                        return HttpResponseRedirect(reverse("stock_web:newinv", args=[form.cleaned_data["reagent"].pk]))
                    elif form.cleaned_data["reagent"].recipe is not None:
                        return HttpResponseRedirect(reverse("stock_web:createnewsol", args=[form.cleaned_data["reagent"].recipe.pk]))
        else:
            form = form()
    else:
        item=Reagents.objects.get(pk=int(pk))
        if item.recipe is not None:
            return HttpResponseRedirect(reverse("stock_web:createnewsol", args=[item.recipe_id]))
        title=["Enter Delivery Details - {} {}".format(item, "- " + item.cat_no if item.cat_no is not None else"")]
        template="stock_web/newinvform.html"
        if item.track_vol==False:
            form=NewInvForm
        elif item.track_vol==True:
            form=NewProbeForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST,instance=item)
                if form.is_valid():
                    ids=Inventory.create(form.cleaned_data, httprequest.user)
                    message=[]
                    if item.track_vol==False:
                        quant=form.cleaned_data["reagent"].count_no+int(form.data["num_rec"])
                    elif item.track_vol==True:
                        quant=form.cleaned_data["reagent"].count_no+int(form.data["vol_rec"])
                    if quant<form.cleaned_data["reagent"].min_count:
                        if form.cleaned_data["reagent"].recipe is not None:
                            make="made"
                        else:
                            make="ordered"
                        message+=["Current stock level for {} is {}. Minimum quantity is {}. Check if more needs to be {}".format(form.cleaned_data["reagent"].name,
                                                                                                                                       quant,
                                                                                                                                       form.cleaned_data["reagent"].min_count,
                                                                                                                                       make)]
                    if item.recipe is None:
                        items=Inventory.objects.filter(reagent=form.cleaned_data["reagent"].id, lot_no=form.cleaned_data["lot_no"], date_rec=form.cleaned_data["date_rec"], val_id__gte=0)
                        # import pdb; pdb.set_trace()
                        if len(items)>0:
                             message+=["THIS ITEM IS VALIDATED. RUN {}".format(items[0].val.val_run)]
                        else:
                             message+=["THIS ITEM IS NOT VALIDATED"]
                        if item.track_vol==False:
                            messages.info(httprequest, "{}x {} added".format(form.data["num_rec"],
                                                                              form.cleaned_data["reagent"]))
                        elif item.track_vol==True:
                            messages.info(httprequest, "1x {}µl of {} added".format(form.data["vol_rec"],
                                                                              form.cleaned_data["reagent"]))
                    if message!=[]:
                        messages.success(httprequest," ".join(message))
                    messages.info(httprequest, "STOCK NUMBERS:")
                    for ID in ids:
                        messages.info(httprequest, ID)
                    return HttpResponseRedirect(reverse("stock_web:newinv",args=["_"]))
        else:
            form = form(initial = {"supplier":item.supplier_def,
                                   "reagent":item,
                                   "date_rec":datetime.date.today})
    submiturl = reverse("stock_web:newinv",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    if httprequest.user.is_staff:
        active="new"
    else:
        active="New Inventory Item"
    return render(httprequest, template, {"header":title, "form": form, "toolbar": _toolbar(httprequest, active=active), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_logged_in, login_url=LOGINURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def createnewsol(httprequest, pk):
    recipe = Recipe.objects.get(pk=int(pk))
    title="Select Items for Recipe: {}".format(recipe)
    form=WitnessForm
    if httprequest.method == "POST":
        if "submit" in httprequest.POST and httprequest.POST["submit"] == "save":
            form = form(httprequest.POST)
            Reagents=[Inventory.objects.get(pk=int(x)).reagent for x in httprequest.POST.getlist("requests")]
            if len(Reagents)>10:
                messages.success(httprequest, "More Than 10 Items Selected. Max Component Limit is 10 Items")
                return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            Reagents_set=set(Reagents)
            vols_used={}
            vol_made=""
            potentials=recipe.liststock()
            potentials.sort(key=attrgetter("is_op"),reverse=True)
            comp_vol=any(p.current_vol is not None for p in potentials)
            witness=None
            try:
                witness=User.objects.get(pk=int(form.data["name"]))
                if witness==httprequest.user:
                    messages.success(httprequest, "YOU MAY NOT USE YOURSELF AS A WITNESS")
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            except ValueError:
                witness=None
            if recipe.track_vol==True:
                vol_made=httprequest.POST.getlist("total_volume")[0]
                if httprequest.POST.getlist("total_volume")==[""]:
                    messages.success(httprequest, "Total Volume Made Not Entered")
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            if comp_vol==True:
                if all(v=="" for v in httprequest.POST.getlist("volume")):
                    messages.success(httprequest, "No Volumes Entered")
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
                #if the first item isn't the track vol, zip gives the first track vol (item #2) to item #1
                #work around was to give everything that's not track vol a hidden 0, but then those all of those aren't checked so ERRORS
                #find fix for allowing intended 0+ticked to be counted but not giving "volumes entered doesn't match tick boxes" error...?
                vols=zip(potentials,httprequest.POST.getlist("volume"))

                for vol in vols:
                    #skips volumes with "a" as a is the value given to hidden volumes so the zip works properly
                    if vol[1]=="a":
                        continue
                    if vol[1]!="":
                        if str(vol[0].pk) not in httprequest.POST.getlist("requests"):
                            messages.success(httprequest, "Selected Checkmarks and Volume Used boxes do not match")
                            return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
                        elif vol[1]!="0":
                            vols_used[str(vol[0].pk)]=vol[1]
                for req in httprequest.POST.getlist("requests"):
                    if req not in vols_used.keys() and Inventory.objects.get(pk=int(req)).current_vol is not None:
                        messages.success(httprequest, "Selected Checkmarks and Volume Used boxes do not match")
                        return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))

                errors=[]
                sum_vol=0
                for item, vol in vols_used.items():
                    invitem=Inventory.objects.get(pk=int(item))
                    sum_vol+=int(vol)
                    if int(invitem.current_vol)-int(vol)<0:
                        errors+=["Reagent {} only has {}µl in the tube. Cannot take {}µl".format(invitem.reagent.name, invitem.current_vol, vol)]
                if errors!=[]:
                    messages.success(httprequest, " ".join(errors))
                    return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
                if recipe.track_vol==True:
                    if int(vol_made)<sum_vol:
                        messages.success(httprequest, "Total Volume of Reagents Used is {}µl. Total Volume made must be at least this volume".format(sum_vol))
                        return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))



            if len(Reagents_set)!=recipe.length():
                if len(Reagents_set)==1:
                    grammar="item was"
                else:
                    grammar="items were"
                messages.success(httprequest, "Only {} of the required items {} selected. This recipe requires {} items".format(len(Reagents_set),
                                                                                                          grammar,
                                                                                                          recipe.length()))
                return HttpResponseRedirect(reverse("stock_web:createnewsol",args=[pk]))
            un_open=[]
            for item in [Inventory.objects.get(pk=int(x)) for x in httprequest.POST.getlist("requests")]:
                if item.is_op==False:
                    un_open+=["Reagent {} was not previously open. It has now been marked as open on its date received".format(item)]
            if un_open!=[]:
                messages.success(httprequest," ".join(un_open))
            sol=Solutions.create(recipe, [int(x) for x in httprequest.POST.getlist("requests") if x.isdigit()], vols_used, vol_made, httprequest.user, witness)
            pk=Inventory.objects.get(internal__batch_number=sol[0]).pk
            return HttpResponseRedirect(reverse("stock_web:item", args=[pk]))
        else:
            HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        form=form()
        values=[]
        inv_ids = []
        checked = []
        VOL=[]
        potentials=recipe.liststock()
        potentials.sort(key=attrgetter("is_op"),reverse=True)
        comp_vol=any(p.current_vol is not None for p in potentials)
        if comp_vol==False:
            vol=False
            headings = ["Reagent Name", "Supplier", "Expiry Date", "Stock Number", "Lot Number", "Date Open", "Validation Run", "Select", ""]
        elif comp_vol==True:
            vol=True
            headings = ["Reagent Name", "Supplier", "Expiry Date", "Stock Number", "Lot Number", "Date Open", "Current Volume", "Validation Run", "Select", "Volume used (µl)", ""]
        for p in potentials:
            #temp used so that can make array for that item, then insert if it's volume is tracked
            temp=[p.reagent.name,
                  p.supplier.name,
                  p.date_exp,
                  p.internal.batch_number,
                  p.lot_no,
                  p.date_op if p.date_op is not None else "NOT OPEN",
                  p.val.val_run if p.val is not None else ""]
            if vol==True:
                temp.insert(-1, "{}µl".format(p.current_vol) if p.current_vol is not None else "N/A")
            values.append(temp)
            inv_ids.append(p.id)
            checked.append("")
            VOL.append(p.current_vol is not None)
        context = { "headings": headings,
                    "body": zip(values, inv_ids, checked, VOL),
                    "url": reverse("stock_web:createnewsol", args=[pk]),
                    "toolbar": _toolbar(httprequest),
                    "total":recipe.track_vol,
                    "identifier":title,
                    "form":form,
                    "cancelurl": reverse("stock_web:newinv",args=["_"]),
                  }
        return render(httprequest, "stock_web/populatesol.html", context)

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newreagent(httprequest):
    form=NewReagentForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Reagents.create(form.cleaned_data)
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newreagent"))
    else:
        form = form()
    submiturl = reverse("stock_web:newreagent")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Reagent Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newsup(httprequest):
    form=NewSupForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Suppliers.create(form.cleaned_data["name"])
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newsup"))
    else:
        form = form()
    submiturl = reverse("stock_web:newsup")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Supplier Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newproj(httprequest):
    form=NewProjForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Projects.create(form.cleaned_data["name"])
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newproj"))
    else:
        form = form()
    submiturl = reverse("stock_web:newproj")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Project Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newstore(httprequest):
    form=NewStoreForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                Storage.create(form.cleaned_data["name"])
                messages.info(httprequest, "Location: {} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newstore"))
    else:
        form = form()
    submiturl = reverse("stock_web:newstore")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Storage Location Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})


@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def newrecipe(httprequest):
    form=NewRecipeForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():

                Recipe.create(form.cleaned_data)
                messages.info(httprequest, "{} Added".format(form.cleaned_data["name"]))
                return HttpResponseRedirect(reverse("stock_web:newrecipe"))
    else:
        form = form()
    submiturl = reverse("stock_web:newrecipe")
    cancelurl = reverse("stock_web:listinv")
    return render(httprequest, "stock_web/form.html", {"header":["New Recipe Input"], "form": form, "toolbar": _toolbar(httprequest, active="new"), "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activsup(httprequest):
    header = ["Select Supplier To Toggle Active State - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditSupForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   message="Supplier {} Has Been Deactivated".format(form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    message="Supplier {} Has Been Reactivated".format(form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activsup"))
    else:
        form = form()

    submiturl = reverse("stock_web:activsup")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activproj(httprequest):
    header = ["Select Project To Toggle Active State - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditProjForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   message="Project {} Has Been Deactivated".format(form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    message="Project {} Has Been Reactivated".format(form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activproj"))
    else:
        form = form()

    submiturl = reverse("stock_web:activproj")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activstore(httprequest):
    header = ["Select Storage Location To Toggle Active State - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditStoreForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   message="Location {} Has Been Deactivated".format(form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    message="Location {} Has Been Reactivated".format(form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activproj"))
    else:
        form = form()

    submiturl = reverse("stock_web:activstore")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})



@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def activreag(httprequest):
    header = ["Select Reagent To Toggle Active State - THIS WILL NOT AFFECT EXISTING ITEMS"]
    form=EditReagForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                if form.cleaned_data["name"].recipe is None:
                    word="Reagent"
                else:
                    word="Recipe"
                if form.cleaned_data["name"].is_active==True:
                   form.cleaned_data["name"].is_active=False
                   message="{} {} Has Been Deactivated".format(word, form.cleaned_data["name"].name)
                else:
                    form.cleaned_data["name"].is_active=True
                    message="{} {} Has Been Reactivated".format(word, form.cleaned_data["name"].name)
                form.cleaned_data["name"].save()
                messages.success(httprequest, message)
                return HttpResponseRedirect(reverse("stock_web:activreag"))
    else:
        form = form()

    submiturl = reverse("stock_web:activreag")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl, "active":"admin"})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def changemin(httprequest, pk):
    submiturl = reverse("stock_web:changemin",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        header = "Select Reagent to Change Minimum Stock Level"
        form=ChangeMinForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:changemin", args=[form.cleaned_data["name"].pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/undoform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item=Reagents.objects.get(pk=int(pk))
        if item.track_vol==True:
            header = ["Select New Default Minimum Stock Level (in µl) for {}".format(item)]
        else:
            header = ["Select New Default Minimum Stock Level for {}".format(item)]
        form=ChangeMinForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:changedef", args=["_"]))
            else:
                form = form(httprequest.POST, initial = {"old":item.min_count})
                if form.is_valid():
                    item.min_count=form.cleaned_data["number"]
                    item.save()
                    messages.success(httprequest, "Minimum Stock Number for {} has changed to {}{}".format(item,form.cleaned_data["number"], "µl" if item.track_vol==True else ""))
                    return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = form(initial = {"old":item.min_count})
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def changedef(httprequest, pk):
    submiturl = reverse("stock_web:changedef",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        header = "Select Reagent to Change Default Supplier - THIS WILL NOT AFFECT EXISTING ITEMS"
        form=ChangeDefForm1
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:changedef", args=[form.cleaned_data["name"].pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/undoform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item=Reagents.objects.get(pk=int(pk))
        header = ["Select New Default Supplier for {} - THIS WILL NOT AFFECT EXISTING ITEMS".format(item)]
        form=ChangeDefForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:changedef", args=["_"]))
            else:
                form = form(httprequest.POST, initial = {"old":item.supplier_def})
                if form.is_valid():
                    item.supplier_def=form.cleaned_data["supplier_def"]
                    item.save()
                    messages.success(httprequest, "Default supplier for {} has changed to {}".format(item,form.cleaned_data["supplier_def"].name))
                    return HttpResponseRedirect(reverse("stock_web:listinv"))
        else:
            form = form(initial = {"supplier_def":item.supplier_def,
                                   "old":item.supplier_def})
    return render(httprequest, "stock_web/form.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def removesup(httprequest):
    header = "Select Supplier To Remove"
    form=RemoveSupForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                form.cleaned_data["supplier"].delete()
                messages.success(httprequest, "Supplier {} Has Been Deleted".format(form.cleaned_data["supplier"].name))
                return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        form = form()
    submiturl = reverse("stock_web:removesup")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    return render(httprequest, "stock_web/undoform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def removeproj(httprequest):
    header = "Select Project To Remove"
    form=RemoveProjForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                form.cleaned_data["project"].delete()
                messages.success(httprequest, "Project {} Has Been Deleted".format(form.cleaned_data["project"].name))
                return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        form = form()
    submiturl = reverse("stock_web:removeproj")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    return render(httprequest, "stock_web/undoform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def removestore(httprequest):
    header = "Select Storage Location To Remove"
    form=RemoveStoreForm
    if httprequest.method=="POST":
        if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
            return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
        else:
            form = form(httprequest.POST)
            if form.is_valid():
                form.cleaned_data["storage"].delete()
                messages.success(httprequest, "Storage Location {} Has Been Deleted".format(form.cleaned_data["storage"].name))
                return HttpResponseRedirect(reverse("stock_web:listinv"))
    else:
        form = form()
    submiturl = reverse("stock_web:removestore")
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    return render(httprequest, "stock_web/undoform.html", {"header": header, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def editinv(httprequest, pk):
    submiturl = reverse("stock_web:editinv",args=[pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")
    if pk=="_":
        title="***WARNING - ONLY TO BE USED TO CORRECT DATA ENTRY ERRORS. IT MAY NOT BE POSSIBLE TO UNDO CHANGES MADE HERE***"
        form=EditInvForm
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "search":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    return HttpResponseRedirect(reverse("stock_web:editinv", args=[Inventory.objects.get(internal=Internal.objects.get(batch_number=form.cleaned_data["item"])).pk]))
        else:
            form = form()
        return render(httprequest, "stock_web/undoform.html", {"header":title, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
    else:
        item = Inventory.objects.get(pk=int(pk))
        if item.reagent.track_vol==True:
            return render(httprequest, "stock_web/list_item.html", _vol_context(httprequest, item, "undo"))
        else:
            return render(httprequest, "stock_web/list_item.html", _item_context(httprequest, item, "undo"))

@user_passes_test(is_admin, login_url=UNAUTHURL)
@user_passes_test(no_reset, login_url=RESETURL, redirect_field_name=None)
def undoitem(httprequest, task, pk):
    item=Inventory.objects.get(pk=int(pk))
    submiturl = reverse("stock_web:undoitem",args=[task, pk])
    cancelurl = reverse("stock_web:listinv")
    toolbar = _toolbar(httprequest, active="Edit Data")

    if task in ["delete", "unopen", "reopen","unuse"]:
        form = DeleteForm
        title=["ARE YOU SURE YOU WANT TO {} ITEM {} - {} {}".format(task.upper(), item.internal, item.reagent,"({}µl use)".format(item.last_usage.used) if task=="unuse" else "")]
        if task=="unopen" and item.reagent.track_vol==True and item.current_vol!=item.vol_rec:
            title+=["THIS WILL REMOVE ALL USES OF THIS REAGENT AND SET ITS VOLUME BACK TO ITS VOLUME RECEIVED"]
        #pdb.set_trace()
        if httprequest.method=="POST":

            if "submit" not in httprequest.POST:
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                if form.is_valid():
                    if form.cleaned_data["sure"]==True:
                        with transaction.atomic():
                            if task=="reopen":
                                item.finished=0
                                item.fin_user=None
                                item.fin_text=None
                                item.reagent.save()
                                item.save()
                            elif task=="unopen":
                                sols=Solutions.objects.filter(Q(comp1=item)|Q(comp2=item)|Q(comp3=item)|
                                                              Q(comp4=item)|Q(comp5=item)|Q(comp6=item)|
                                                              Q(comp7=item)|Q(comp8=item)|Q(comp9=item)|
                                                              Q(comp10=item))
                                if len(sols)!=0:
                                    messages.info(httprequest, "UNABLE TO UNOPEN ITEM AS IT'S USED IN THE FOLLOWING SOLUTION(S)")
                                    for sol in sols:
                                        #pdb.set_trace()
                                        messages.info(httprequest,Inventory.objects.get(sol=sol))
                                    return HttpResponseRedirect(reverse("stock_web:undoitem", args=[task,pk]))
                                item.date_op=None
                                item.is_op=0
                                item.op_user_id=None
                                item.project_used=None
                                if item.reagent.track_vol:
                                    item.reagent.count_no+=sum([x.used for x in VolUsage.objects.filter(item=item)])
                                    item.reagent.save()
                                    item.last_usage=None
                                    item.current_vol=item.vol_rec
                                    item.save()
                                    VolUsage.objects.filter(item=item).delete()
                                else:
                                    item.reagent.count_no+=1
                                    item.reagent.save()
                                item.save()
                            if task=="unuse" and item.reagent.track_vol==True:
                                uses=VolUsage.objects.filter(item=item).order_by("id").reverse()
                                use=item.last_usage
                                if use.sol is not None:
                                    messages.success(httprequest, "You cannot undo this usage as it was part of solution {}. PLEASE EDIT THIS SOLUTION TO UNDO THIS USAGE".format(Inventory.objects.get(sol=use.sol)))
                                    return HttpResponseRedirect(reverse("stock_web:undoitem", args=[task,pk]))
                                item.current_vol+=use.used
                                item.reagent.count_no+=use.used
                                if len(uses)>1:
                                    item.last_usage=uses[1]
                                else:
                                    item.last_usage=None
                                item.reagent.save()
                                if item.finished==True:
                                    item.finished=False
                                    item.date_fin=None
                                    item.fin_user=None
                                    item.fin_text=None
                                item.save()
                                use.delete()
                            if task=="delete":
                                if item.reagent.track_vol==False:
                                    item.reagent.count_no-=1
                                else:
                                    item.reagent.count_no-=item.current_vol
                                item.reagent.save()
                                item.delete()
                                if item.sol is not None:
                                    sol=item.sol
                                    if item.reagent.track_vol==True:
                                        for i in range(sol.recipe.length()):
                                            comp=eval('sol.comp{}'.format(i+1))
                                            uses=VolUsage.objects.filter(item=comp).order_by("id").reverse()
                                            last_use=comp.last_usage
                                            if last_use==uses.get(sol=sol):
                                                try:
                                                    comp.last_usage=uses[1]
                                                except:
                                                    comp.last_usage=None


                                            comp.current_vol+=uses.get(sol=sol).used
                                            comp.reagent.count_no+=uses.get(sol=sol).used
                                            if comp.finished==True:
                                                comp.finished=False
                                                comp.date_fin=None
                                                comp.fin_user=None
                                                comp.fin_text=None
                                            comp.save()
                                            comp.reagent.save()
                                            uses.get(sol=sol).delete()

                                        message="ITEM DELETED AND VOLUMES OF COMPONENTS USED FOR SOLUTION HAVE BEEN REPLACED"
                                    else:
                                        message='Item {} has been deleted!'.format(item)
                                    sol.delete()
                                else:
                                    message='Item {} has been deleted!'.format(item)
                                messages.success(httprequest,message)
                                return HttpResponseRedirect(reverse("stock_web:listinv"))



                    return HttpResponseRedirect(reverse("stock_web:editinv", args=[pk]))
        else:
            form = form()
    elif task=="unval":
        form=UnValForm
        title=["ARE YOU SURE YOU WANT TO UN-VALIDATE ITEM {} - {}".format(item.internal, item.reagent)]
        if httprequest.method=="POST":
            if "submit" not in httprequest.POST or httprequest.POST["submit"] != "save":
                return HttpResponseRedirect(httprequest.session["referer"] if ("referer" in httprequest.session) else reverse("stock_web:listinv"))
            else:
                form = form(httprequest.POST)
                form.fields["all_type"].choices = [(0,"NO"),
                                          (1,"YES - Only For This Reagent"),
                                          (2,"YES - All Items On {}, Regardless of Reagent".format(item.val))]

                if form.is_valid():
                    if form.cleaned_data["sure"]==True:
                        current_val_id=item.val_id
                        items=[]
                        if int(form.cleaned_data["all_type"])==0:
                            item.val_id=None
                            item.save()
                        elif int(form.cleaned_data["all_type"])==1:
                            items=Inventory.objects.filter(val_id=current_val_id, reagent_id=item.reagent_id)
                            messages.success(httprequest, "UNVALIDATED: {}".format(", ".join(test.internal.batch_number for test in items)))
                            items.update(val_id=None)
                        elif int(form.cleaned_data["all_type"])==2:
                            items=Inventory.objects.filter(val_id=current_val_id)
                            messages.success(httprequest, "UNVALIDATED: {}".format(", ".join(test.internal.batch_number for test in items)))
                            items.update(val_id=None)
                        if len(Inventory.objects.filter(val_id=current_val_id))==0:
                            Validation.objects.get(pk=current_val_id).delete()
                    return HttpResponseRedirect(reverse("stock_web:editinv", args=[pk]))
        else:
            #pdb.set_trace()
            form = form()
            form.fields["all_type"].choices = [(0,"NO"),
                                          (1,"YES - Only For This Reagent"),
                                          (2,"YES - All Items On {}, Regardless of Reagent".format(item.val))]
    return render(httprequest, "stock_web/form.html", {"header":title, "form": form, "toolbar": toolbar, "submiturl": submiturl, "cancelurl": cancelurl})
